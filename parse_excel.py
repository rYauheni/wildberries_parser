import openpyxl


def extract_ids_from_excel(file_path: str) -> list:
    ids = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                ids.append(row[0])

    except Exception as e:
        print(f"Произошла ошибка: {e}")

    return ids

