import openpyxl

from exceptions.exceptions import SourceError
from logger_utils.logger_utils import logger
from product_id_extract_services.file_provider import FileProvider
from product_id_extract_services.product_id_extract_service import ProductIDExtractService


class ExcelProductIDExtractService(ProductIDExtractService):
    ids: list[int]

    def get_ids(self) -> list:

        # You cannot pass file_path as an argument to this method,
        # because not every ProductIDExtractService (for example service for database)
        # will have the file_path parameter.
        # Compliance with the LSP.
        file_path = FileProvider().get_file_path()
        logger.info(f'ID source is excel file: {file_path}')
        self.ids = []
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
                if row[0] is not None:
                    self.ids.append(row[0])

        except Exception:
            raise SourceError

        return self.ids
