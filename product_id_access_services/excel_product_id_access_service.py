import sys

import openpyxl

from exceptions.exceptions import FileError
from product_id_access_services.product_id_access_service import ProductIDAccessService
from settings import DEFAULT_EXCEL_FILE_PATH


class ExcelProductIDAccessService(ProductIDAccessService):
    ids: list[int]
    source: str

    def get_excel_file_path(self) -> str:
        args = sys.argv
        if '--excel_file_path' not in args:
            excel_file_path = DEFAULT_EXCEL_FILE_PATH
        else:
            excel_file_path_index = args.index("--excel_file_path") + 1
            excel_file_path = args[excel_file_path_index]
        self.source = excel_file_path
        return excel_file_path

    def get_ids(self) -> list:
        file_path = self.get_excel_file_path()
        self.ids = []
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
                if row[0] is not None:
                    self.ids.append(row[0])

        except Exception:
            raise FileError

        return self.ids
