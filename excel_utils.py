import sys

import openpyxl

from settings import DEFAULT_EXCEL_FILE_PATH
from exceptions.exceptions import FileError


def get_excel_file_path() -> str:
    args = sys.argv
    if '--excel_file_path' not in args:
        excel_file_path = DEFAULT_EXCEL_FILE_PATH
    else:
        excel_file_path_index = args.index("--excel_file_path") + 1
        excel_file_path = args[excel_file_path_index]

    return excel_file_path


def extract_ids_from_excel(file_path: str) -> list:
    ids = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0] is not None:
                ids.append(row[0])

    except Exception:
        raise FileError

    return ids

